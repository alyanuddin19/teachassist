import csv
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from uuid import uuid4

import pandas as pd
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook

from .config import get_mapping_config


UPLOAD_DIR = Path(os.getenv("COLUMN_MAPPING_UPLOAD_DIR", Path("uploads") / "column_mapping"))
OUTPUT_DIR = Path(os.getenv("COLUMN_MAPPING_OUTPUT_DIR", Path("uploads") / "column_mapping_outputs"))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class WorkbookFile:
    path: Path
    original_name: str


def save_upload(upload: UploadFile) -> WorkbookFile:
    config = get_mapping_config()
    suffix = Path(upload.filename or "").suffix.lower()
    if suffix not in set(config.get("supported_extensions", [])):
        raise HTTPException(status_code=400, detail="Please upload a supported Excel or CSV file.")

    max_bytes = int(config.get("max_upload_mb", 15)) * 1024 * 1024
    path = UPLOAD_DIR / f"{uuid4().hex}{suffix}"
    size = 0
    with path.open("wb") as target:
        while True:
            chunk = upload.file.read(1024 * 1024)
            if not chunk:
                break
            size += len(chunk)
            if size > max_bytes:
                path.unlink(missing_ok=True)
                raise HTTPException(status_code=400, detail="File is too large.")
            target.write(chunk)

    if size == 0:
        path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    return WorkbookFile(path=path, original_name=upload.filename or "workbook")


def is_html_spreadsheet(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            start = handle.read(256).lstrip().lower()
        return start.startswith(b"<html") or start.startswith(b"<!doctype html")
    except OSError:
        return False


def load_any_workbook(path: Path, display_name: str | None = None):
    suffix = path.suffix.lower()
    try:
        if suffix == ".csv":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = safe_sheet_title(Path(display_name or path.name).stem, "Sheet1")
            with path.open(newline="", encoding="utf-8", errors="ignore") as handle:
                for row in csv.reader(handle):
                    sheet.append(row)
            return workbook

        if suffix == ".xls" and is_html_spreadsheet(path):
            return html_spreadsheet_to_workbook(path, display_name)

        if suffix == ".xls":
            sheets = pd.read_excel(path, sheet_name=None, dtype=object, header=None)
            workbook = Workbook()
            workbook.remove(workbook.active)
            for index, (sheet_name, dataframe) in enumerate(sheets.items(), start=1):
                sheet = workbook.create_sheet(safe_sheet_title(sheet_name, f"Sheet{index}"))
                for row in dataframe.fillna("").values.tolist():
                    sheet.append(row)
            return workbook

        return load_workbook(path)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read workbook. It may be corrupted or password protected.") from exc


def html_spreadsheet_to_workbook(path: Path, display_name: str | None = None):
    try:
        tables = pd.read_html(path, header=None, keep_default_na=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Could not read this Excel export.") from exc
    if not tables:
        raise HTTPException(status_code=400, detail="No readable table was found in this file.")

    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, dataframe in enumerate(tables, start=1):
        fallback = Path(display_name or path.name).stem if len(tables) == 1 else f"Table {index}"
        sheet = workbook.create_sheet(safe_sheet_title(fallback, f"Sheet{index}"))
        for row in dataframe.fillna("").values.tolist():
            sheet.append(row)
    return workbook


def safe_sheet_title(value: object, fallback: str) -> str:
    title = "".join(ch for ch in str(value or fallback) if ch not in r"[]:*?/\\").strip()[:31]
    return title or fallback


def discover_sheets(file: WorkbookFile) -> list[str]:
    workbook = load_any_workbook(file.path, file.original_name)
    return workbook.sheetnames


def get_sheet(workbook, sheet_name: str):
    if not sheet_name or sheet_name not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail="Please select a valid sheet.")
    return workbook[sheet_name]


def row_values(sheet, row_number: int) -> list[object]:
    return [sheet.cell(row=row_number, column=col).value for col in range(1, sheet.max_column + 1)]


def is_blank_row(sheet, row_number: int) -> bool:
    return all(value in (None, "") for value in row_values(sheet, row_number))


def normalized_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def is_student_identifier(value: object) -> bool:
    text = normalized_text(value)
    if not text:
        return False
    return any(term in text for term in [
        "registration",
        "reg no",
        "reg number",
        "roll no",
        "roll number",
        "student id",
        "student name",
    ]) or text in {"name", "student", "full name"}


def detect_student_header_row(sheet, max_scan: int | None = None) -> int | None:
    scan_limit = min(sheet.max_row, max_scan or 40)
    best_row = None
    best_score = 0
    for row in range(1, scan_limit + 1):
        values = row_values(sheet, row)
        identifiers = [value for value in values if is_student_identifier(value)]
        if not identifiers:
            continue
        text_count = sum(
            1
            for value in values
            if value not in (None, "") and any(ch.isalpha() for ch in str(value))
        )
        score = len(identifiers) * 10 + text_count
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def detect_header_row(sheet, configured_row: int | None = None) -> int:
    if configured_row and configured_row > 0:
        return configured_row
    student_header = detect_student_header_row(sheet)
    if student_header:
        return student_header

    best_row = 0
    best_score = -1
    max_scan = min(sheet.max_row, 25)
    for row in range(1, max_scan + 1):
        values = row_values(sheet, row)
        non_empty = [str(value).strip() for value in values if value not in (None, "")]
        text_count = sum(1 for value in non_empty if any(ch.isalpha() for ch in value))
        unique_count = len(set(non_empty))
        score = text_count * 3 + unique_count + len(non_empty)
        if score > best_score:
            best_score = score
            best_row = row

    if best_row <= 0 or best_score <= 0:
        raise HTTPException(status_code=400, detail="Could not find a header row.")
    return best_row


def detect_data_start_row(sheet, header_row: int, configured_row: int | None = None) -> int:
    if configured_row and configured_row > 0:
        return configured_row
    for row in range(header_row + 1, sheet.max_row + 1):
        if not is_blank_row(sheet, row):
            return row
    return header_row + 1


def extract_headers(sheet, header_row: int) -> list[dict]:
    base_headers = []
    for column in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=column).value
        label = header_label_for_column(sheet, header_row, column)
        base_headers.append((label, raw, column))

    label_counts: dict[str, int] = {}
    for label, _, _ in base_headers:
        label_counts[label] = label_counts.get(label, 0) + 1

    headers = []
    seen: dict[str, int] = {}
    for label, raw, column in base_headers:
        seen[label] = seen.get(label, 0) + 1
        display_label = display_header_label(label, seen[label], label_counts[label])
        headers.append({
            "name": display_label,
            "column_index": column,
            "duplicate_index": seen[label],
            "is_empty": raw in (None, ""),
        })
    return headers


def header_label_for_column(sheet, header_row: int, column: int) -> str:
    raw = sheet.cell(row=header_row, column=column).value
    if raw not in (None, ""):
        return str(raw).strip()

    parts: list[str] = []
    for row in range(header_row - 1, 0, -1):
        if is_metadata_row(sheet, row):
            continue
        text = inherited_header_text(sheet, row, column)
        if not text:
            continue
        if text.lower() in {part.lower() for part in parts}:
            continue
        parts.insert(0, text)
        if len(parts) >= 3:
            break
    return " ".join(parts) if parts else f"Column {column}"


def display_header_label(label: str, occurrence: int, total: int) -> str:
    if total <= 1:
        return label
    if is_repeatable_assessment_label(label) and not has_explicit_assessment_number(label):
        return f"{label} {occurrence}"
    return label if occurrence == 1 else f"{label} ({occurrence})"


def is_repeatable_assessment_label(label: str) -> bool:
    text = normalized_text(label)
    return "assignment" in text or "quiz" in text


def has_explicit_assessment_number(label: str) -> bool:
    text = normalized_text(label)
    return bool(re.search(r"\b(assignment|quiz)\s*\d+\b", text))


def is_metadata_row(sheet, row: int) -> bool:
    values = [normalized_text(value) for value in row_values(sheet, row) if value not in (None, "")]
    if not values:
        return False
    marker_text = " ".join(values[:3])
    if any(marker in marker_text for marker in ["% weight", "weight", "marks", "max marks", "total marks", "kpi"]):
        return True
    numeric_count = sum(1 for value in values if re.fullmatch(r"\d+(\.\d+)?%?", value))
    return numeric_count >= max(3, len(values) - 1)


def inherited_header_text(sheet, row: int, column: int) -> str:
    value = sheet.cell(row=row, column=column).value
    if value not in (None, ""):
        return str(value).strip()

    # Many Excel exports look merged visually, but only store the group label
    # in the left-most cell. Reuse that label for columns under the same group.
    for left_col in range(column - 1, 0, -1):
        left_value = sheet.cell(row=row, column=left_col).value
        if left_value in (None, ""):
            continue
        return str(left_value).strip()
    return ""


def generated_output_path() -> Path:
    return OUTPUT_DIR / f"mapped_{uuid4().hex}.xlsx"


def cleanup_paths(paths: list[Path]) -> None:
    for path in paths:
        try:
            if path.exists() and path.is_file():
                path.unlink()
        except OSError:
            pass
