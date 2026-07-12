from copy import copy
from datetime import date
from pathlib import Path

from fastapi import HTTPException
from openpyxl.formula.translate import Translator

from .workbook_service import (
    detect_data_start_row,
    detect_header_row,
    extract_headers,
    generated_output_path,
    get_sheet,
    load_any_workbook,
)


def validate_mappings(mappings: list[dict], required_targets: list[str], allow_duplicates: bool) -> None:
    selected_targets = [item.get("target_column") for item in mappings if item.get("target_column")]
    if not allow_duplicates and len(selected_targets) != len(set(selected_targets)):
        raise HTTPException(status_code=400, detail="One target column is selected more than once.")

    missing = [target for target in required_targets if target not in selected_targets]
    if missing:
        raise HTTPException(status_code=400, detail=f"Required target columns are unmapped: {', '.join(missing)}")


def apply_mapping(
    source_path: Path,
    source_name: str,
    target_path: Path,
    target_name: str,
    source_sheet_name: str,
    target_sheet_name: str,
    mappings: list[dict],
    required_targets: list[str],
    allow_duplicates: bool,
    source_header_row: int | None = None,
    target_header_row: int | None = None,
    source_data_start_row: int | None = None,
    target_data_start_row: int | None = None,
    target_fields: list[dict] | None = None,
    manual_values: dict[str, str] | None = None,
) -> Path:
    validate_mappings(mappings, required_targets, allow_duplicates)

    source_workbook = load_any_workbook(source_path, source_name)
    target_workbook = load_any_workbook(target_path, target_name)
    source_sheet = get_sheet(source_workbook, source_sheet_name)
    target_sheet = get_sheet(target_workbook, target_sheet_name)

    source_header = detect_header_row(source_sheet, source_header_row)
    target_header = detect_header_row(target_sheet, target_header_row)
    source_start = detect_data_start_row(source_sheet, source_header, source_data_start_row)
    target_start = detect_data_start_row(target_sheet, target_header, target_data_start_row)

    source_headers = {header["name"]: header["column_index"] for header in extract_headers(source_sheet, source_header)}
    target_headers = {header["name"]: header["column_index"] for header in extract_headers(target_sheet, target_header)}

    column_pairs: list[tuple[int, int, str]] = []
    mapped_source_names: set[str] = set()
    for mapping in mappings:
        source_column = mapping.get("source_column")
        target_column = mapping.get("target_column")
        if not source_column or not target_column:
            continue
        if source_column not in source_headers or target_column not in target_headers:
            raise HTTPException(status_code=400, detail="Invalid mapping selection.")
        column_pairs.append((source_headers[source_column], target_headers[target_column], target_column))
        mapped_source_names.add(source_column)

    template_row = target_start if target_start <= target_sheet.max_row else max(target_header, 1)
    write_row = target_start
    source_data_columns = list(source_headers.values())
    for source_row in range(source_start, source_sheet.max_row + 1):
        columns_to_check = [source_col for source_col, _, _ in column_pairs] or source_data_columns
        if all(source_sheet.cell(row=source_row, column=source_col).value in (None, "") for source_col in columns_to_check):
            continue
        prepare_target_row(target_sheet, template_row, write_row)
        for source_col, target_col, _ in column_pairs:
            target_sheet.cell(row=write_row, column=target_col).value = source_sheet.cell(row=source_row, column=source_col).value
        apply_missing_rules(
            target_sheet,
            write_row,
            target_start,
            target_headers,
            target_fields or [],
            {name for _, _, name in column_pairs},
            manual_values or {},
        )
        write_row += 1

    add_unmapped_source_sheet(
        target_workbook,
        source_sheet,
        source_headers,
        mapped_source_names,
        source_start,
    )

    output = generated_output_path()
    target_workbook.save(output)
    return output


def apply_missing_rules(sheet, row: int, data_start_row: int, target_headers: dict[str, int], fields: list[dict], mapped_targets: set[str], manual_values: dict[str, str]) -> None:
    for field in fields:
        name = field.get("display_name")
        if not name or name in mapped_targets or name not in target_headers:
            continue
        col = target_headers[name]
        rule = field.get("missing_value_rule") or "blank"
        cell = sheet.cell(row=row, column=col)
        if rule == "blank":
            cell.value = None
        elif rule == "zero":
            cell.value = 0
        elif rule == "fixed":
            cell.value = field.get("default_value") or ""
        elif rule == "manual":
            cell.value = manual_values.get(name, "")
        elif rule == "sequence":
            cell.value = row - data_start_row + 1
        elif rule == "current_date":
            cell.value = date.today().isoformat()
        elif rule == "copy":
            source_name = field.get("default_value") or ""
            source_col = target_headers.get(source_name)
            cell.value = sheet.cell(row=row, column=source_col).value if source_col else None
        elif rule == "formula":
            formula = field.get("formula_definition") or field.get("default_value") or ""
            cell.value = formula if str(formula).startswith("=") else None


def prepare_target_row(sheet, template_row: int, target_row: int) -> None:
    if template_row == target_row:
        return
    if template_row <= sheet.max_row:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[template_row].height
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(row=template_row, column=column)
            target = sheet.cell(row=target_row, column=column)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
                target.protection = copy(source.protection)
            if isinstance(source.value, str) and source.value.startswith("="):
                try:
                    target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)
                except Exception:
                    target.value = source.value


def add_unmapped_source_sheet(workbook, source_sheet, source_headers: dict[str, int], mapped_source_names: set[str], source_start: int) -> None:
    unmapped = [(name, column) for name, column in source_headers.items() if name not in mapped_source_names]
    if not unmapped:
        return
    sheet_name = "Unmapped Columns"
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)
    for col_index, (name, _) in enumerate(unmapped, start=1):
        sheet.cell(row=1, column=col_index).value = name
    write_row = 2
    for source_row in range(source_start, source_sheet.max_row + 1):
        values = [source_sheet.cell(row=source_row, column=column).value for _, column in unmapped]
        if all(value in (None, "") for value in values):
            continue
        for col_index, value in enumerate(values, start=1):
            sheet.cell(row=write_row, column=col_index).value = value
        write_row += 1
