import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.column_mapping.matching_service import ColumnMatchingService
from app.services.column_mapping.transformation_service import apply_mapping, validate_mappings
from app.services.column_mapping.workbook_service import detect_data_start_row, detect_header_row, extract_headers


class ColumnMappingTests(unittest.TestCase):
    def test_matching_layers(self):
        service = ColumnMatchingService()
        cases = [
            ("Registration Number", "Registration Number", "exact"),
            ("registration number", "Registration Number", "normalized"),
            ("Registration-No.", "Registration Number", "normalized"),
            ("Reg No", "Registration Number", "normalized"),
            ("Registratic", "Registration Number", "academic-field"),
            ("Marks Obtained", "Marks Scored", "synonym"),
            ("Mark Scored", "Marks Scored", "semantic"),
        ]
        for source, target, method in cases:
            suggestion = service.suggest([{"name": source}], [{"name": target}])[0]
            self.assertEqual(suggestion["suggested_target_column"], target)
            self.assertEqual(suggestion["matching_method"], method)
        self.assertEqual(
            service.suggest([{"name": "Unknown Field"}], [{"name": "Marks Scored"}])[0]["status"],
            "Unmapped",
        )

    def test_registration_and_name_field_matching(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "Registratic"}, {"name": "Name"}],
            [{"name": "Registration Number"}, {"name": "Student Name"}],
        )

        self.assertEqual(suggestions[0]["suggested_target_column"], "Registration Number")
        self.assertEqual(suggestions[0]["status"], "Mapped")
        self.assertEqual(suggestions[1]["suggested_target_column"], "Student Name")
        self.assertEqual(suggestions[1]["status"], "Mapped")

    def test_semantic_assessment_term_matching(self):
        service = ColumnMatchingService()
        source = [
            {"name": "First Term Q1"},
            {"name": "Final Term Q1"},
            {"name": "Theory End-Term Q2"},
            {"name": "Mid-Term Q2"},
            {"name": "Quiz 01"},
        ]
        targets = [
            {"name": "Mid Term Q1"},
            {"name": "End Term Q1"},
            {"name": "Theory End Term Q2"},
            {"name": "Mid Term Q2"},
            {"name": "Quiz 1"},
        ]
        suggestions = service.suggest(source, targets)

        self.assertEqual(suggestions[0]["suggested_target_column"], "Mid Term Q1")
        self.assertEqual(suggestions[1]["suggested_target_column"], "End Term Q1")
        self.assertEqual(suggestions[2]["suggested_target_column"], "Theory End Term Q2")
        self.assertEqual(suggestions[3]["suggested_target_column"], "Mid Term Q2")
        self.assertEqual(suggestions[4]["suggested_target_column"], "Quiz 1")

    def test_numeric_identifier_prevents_wrong_question_match(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "End Term Q2"}],
            [{"name": "End Term Q1"}, {"name": "End Term Q2"}],
        )

        self.assertEqual(suggestions[0]["suggested_target_column"], "End Term Q2")

    def test_theory_end_term_does_not_map_to_clo_achieved_column(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "Theory End-Term Q2"}],
            [{"name": "CLO Achieved Theory End-Term Q2"}, {"name": "Theory End-Term Q2"}],
        )

        self.assertEqual(suggestions[0]["suggested_target_column"], "Theory End-Term Q2")

        direct = service.suggest(
            [{"name": "Theory End-Term Q2"}],
            [{"name": "CLO Achieved Theory End-Term Q2"}, {"name": "Theory End Term Q2"}],
        )
        self.assertEqual(direct[0]["suggested_target_column"], "Theory End Term Q2")

    def test_clo_context_theory_end_term_prefers_normal_marks_target(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "CLO2 Theory End-Term Q2"}],
            [{"name": "CLO Achieved Theory End-Term Q2"}, {"name": "Theory End Term Q2"}],
        )

        self.assertEqual(suggestions[0]["suggested_target_column"], "Theory End Term Q2")

    def test_unmatched_source_remains_unmapped(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "Unrelated Survey Field"}],
            [{"name": "End Term Q1"}],
        )

        self.assertIsNone(suggestions[0]["suggested_target_column"])
        self.assertEqual(suggestions[0]["status"], "Unmapped")

    def test_global_assignment_resolves_competing_targets(self):
        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": "End Term Q1"}, {"name": "Final Term Q1"}],
            [{"name": "End Term Q1"}, {"name": "End Term Q2"}],
        )

        matched_targets = [item["suggested_target_column"] for item in suggestions if item["suggested_target_column"]]
        self.assertEqual(len(matched_targets), len(set(matched_targets)))
        self.assertIn("End Term Q1", matched_targets)

    def test_duplicate_target_validation(self):
        with self.assertRaises(Exception):
            validate_mappings(
                [
                    {"source_column": "A", "target_column": "X"},
                    {"source_column": "B", "target_column": "X"},
                ],
                [],
                False,
            )

    def test_missing_required_target_validation(self):
        with self.assertRaises(Exception):
            validate_mappings(
                [{"source_column": "A", "target_column": "X"}],
                ["Y"],
                False,
            )

    def test_header_and_data_detection_not_first_row(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Instructions", None])
        sheet.append([None, None])
        sheet.append(["Registration Number", "Marks Scored"])
        sheet.append(["R1", 10])

        header_row = detect_header_row(sheet)
        data_row = detect_data_start_row(sheet, header_row)
        headers = extract_headers(sheet, header_row)

        self.assertEqual(header_row, 3)
        self.assertEqual(data_row, 4)
        self.assertEqual(headers[0]["name"], "Registration Number")

    def test_clo_marksheet_uses_registration_row_and_composite_headers(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["CLO", None, "CLO1", None, None, "CLO2"])
        sheet.append(["Activity", None, "Mid-Term", "Mid-Term", "Quiz 5", "Quiz 4"])
        sheet.append([None, None, "Mid-Term", "Mid-Term", "Quiz 5", "Quiz 4"])
        sheet.append(["% Weight", None, 5, 5, 2, 2])
        sheet.append(["Registration", "Name", None, None, None, None])
        sheet.append(["FA21-001", "Ayesha", 4, 5, 2, 1])

        header_row = detect_header_row(sheet)
        data_row = detect_data_start_row(sheet, header_row)
        headers = extract_headers(sheet, header_row)

        self.assertEqual(header_row, 5)
        self.assertEqual(data_row, 6)
        self.assertEqual(headers[0]["name"], "Registration")
        self.assertEqual(headers[1]["name"], "Name")
        self.assertEqual(headers[2]["name"], "CLO1 Mid-Term")
        self.assertEqual(headers[3]["name"], "CLO1 Mid-Term (2)")
        self.assertEqual(headers[4]["name"], "CLO1 Quiz 5")

    def test_weight_row_does_not_become_assignment_or_quiz_number(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["CLO", None, "CLO1", None, None, None, None])
        sheet.append(["Activity", None, "Assignment", "Assignment", "Quiz 3", "Quiz 5", "Quiz"])
        sheet.append([None, None, "Assignment", "Assignment", "Quiz 3", "Quiz 5", "Quiz"])
        sheet.append(["% Weight", None, 10, 10, 2, 2, 1])
        sheet.append(["Registration", "Name", None, None, None, None, None])
        sheet.append(["S1", "Ayesha", 9, 8, 2, 1, 1])

        header_row = detect_header_row(sheet)
        headers = extract_headers(sheet, header_row)

        self.assertEqual(headers[2]["name"], "CLO1 Assignment 1")
        self.assertEqual(headers[3]["name"], "CLO1 Assignment 2")
        self.assertEqual(headers[4]["name"], "CLO1 Quiz 3")
        self.assertEqual(headers[5]["name"], "CLO1 Quiz 5")
        self.assertEqual(headers[6]["name"], "CLO1 Quiz")

        service = ColumnMatchingService()
        suggestions = service.suggest(
            [{"name": headers[2]["name"]}, {"name": headers[4]["name"]}, {"name": headers[5]["name"]}],
            [{"name": "Assignment 1"}, {"name": "Quiz 3"}, {"name": "Quiz 5"}],
        )

        self.assertEqual(suggestions[0]["suggested_target_column"], "Assignment 1")
        self.assertEqual(suggestions[1]["suggested_target_column"], "Quiz 3")
        self.assertEqual(suggestions[2]["suggested_target_column"], "Quiz 5")

    def test_transform_preserves_template_format_and_formula(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            source_path = temp / "source.xlsx"
            target_path = temp / "target.xlsx"

            source = Workbook()
            source_sheet = source.active
            source_sheet.title = "Source"
            source_sheet.append(["Reg No", "Marks", "Source Only"])
            source_sheet.append(["S1", 8, "Keep 1"])
            source_sheet.append(["S2", 9, "Keep 2"])
            source.save(source_path)

            target = Workbook()
            target_sheet = target.active
            target_sheet.title = "Template"
            target_sheet.append(["Registration Number", "Marks Scored", "Missing Target", "Double Marks"])
            target_sheet.append([None, None, None, "=B2*2"])
            target_sheet["A1"].font = Font(bold=True)
            target_sheet["B2"].fill = PatternFill("solid", fgColor="FFFF00")
            target.save(target_path)

            output = apply_mapping(
                source_path,
                "source.xlsx",
                target_path,
                "target.xlsx",
                "Source",
                "Template",
                [
                    {"source_column": "Reg No", "target_column": "Registration Number"},
                    {"source_column": "Marks", "target_column": "Marks Scored"},
                ],
                ["Registration Number"],
                False,
            )

            result = load_workbook(output, data_only=False)
            sheet = result["Template"]
            self.assertEqual(sheet["A2"].value, "S1")
            self.assertEqual(sheet["B3"].value, 9)
            self.assertIsNone(sheet["C2"].value)
            self.assertEqual(sheet["D3"].value, "=B3*2")
            self.assertEqual(sheet["A1"].font.bold, True)
            self.assertEqual(sheet["B3"].fill.fgColor.rgb, "00FFFF00")
            self.assertIn("Unmapped Columns", result.sheetnames)
            self.assertEqual(result["Unmapped Columns"]["A1"].value, "Source Only")
            self.assertEqual(result["Unmapped Columns"]["A2"].value, "Keep 1")


if __name__ == "__main__":
    unittest.main()
