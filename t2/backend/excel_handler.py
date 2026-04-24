import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)


def build_marksheet_excel(
    sheet_id: int,
    teacher_name: str,
    teacher_email: str,
    department: str,
    section: str,
    total_marks: float,
    selected_options: dict[str, str],
    assessment_totals: dict[str, float],
    assessment_columns: list[str],
    students: list[dict],
) -> str:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Marks Sheet"

    title_fill = PatternFill(fill_type="solid", start_color="1D4ED8", end_color="1D4ED8")
    title_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(fill_type="solid", start_color="DBEAFE", end_color="DBEAFE")

    metadata_rows = [
        ("Teacher Name", teacher_name),
        ("Teacher Email", teacher_email),
        ("Department", department),
        ("Section", section),
        ("Total Marks", total_marks),
        ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]

    for row_index, (label, value) in enumerate(metadata_rows, start=1):
        worksheet.cell(row=row_index, column=1, value=label)
        worksheet.cell(row=row_index, column=2, value=value)
        worksheet.cell(row=row_index, column=1).fill = title_fill
        worksheet.cell(row=row_index, column=1).font = title_font

    header_row = len(metadata_rows) + 2
    subheader_row = header_row + 1
    base_headers = ["Roll Number", "Student Name", "Department", "Section"]

    current_column = 1
    for header in base_headers:
        worksheet.merge_cells(
            start_row=header_row,
            start_column=current_column,
            end_row=subheader_row,
            end_column=current_column,
        )
        cell = worksheet.cell(row=header_row, column=current_column, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_column += 1

    grouped_columns: dict[str, list[tuple[str, float]]] = {}
    for column_label in assessment_columns:
        if " - " in column_label:
            clo, item = column_label.split(" - ", 1)
        else:
            clo, item = column_label, column_label
        grouped_columns.setdefault(clo, []).append((item, float(assessment_totals.get(column_label, 0))))

    assessment_column_lookup: list[str] = []
    for clo, items in grouped_columns.items():
        start_column = current_column
        end_column = current_column + len(items) - 1
        worksheet.merge_cells(
            start_row=header_row,
            start_column=start_column,
            end_row=header_row,
            end_column=end_column,
        )
        clo_total = sum(total for _, total in items)
        clo_cell = worksheet.cell(row=header_row, column=start_column, value=f"{clo} ({clo_total:g})")
        clo_cell.font = Font(bold=True)
        clo_cell.fill = header_fill
        clo_cell.alignment = Alignment(horizontal="center", vertical="center")

        for item, item_total in items:
            subheader_cell = worksheet.cell(
                row=subheader_row,
                column=current_column,
                value=f"{item} ({item_total:g})",
            )
            subheader_cell.font = Font(bold=True)
            subheader_cell.fill = header_fill
            subheader_cell.alignment = Alignment(horizontal="center", vertical="center")
            assessment_column_lookup.append(f"{clo} - {item}")
            current_column += 1

    worksheet.merge_cells(
        start_row=header_row,
        start_column=current_column,
        end_row=subheader_row,
        end_column=current_column,
    )
    total_cell = worksheet.cell(row=header_row, column=current_column, value="Total Marks")
    total_cell.font = Font(bold=True)
    total_cell.fill = header_fill
    total_cell.alignment = Alignment(horizontal="center", vertical="center")

    for student_index, student in enumerate(students, start=subheader_row + 1):
        base_values = [
            student["roll_number"],
            student["full_name"],
            department,
            section,
        ]
        assessment_values = [student["marks"].get(column, "") for column in assessment_column_lookup]
        total_value = sum(
            float(value)
            for value in student["marks"].values()
            if isinstance(value, (int, float))
        )
        tail_values = [total_value]

        for column_index, value in enumerate(base_values + assessment_values + tail_values, start=1):
            worksheet.cell(row=student_index, column=column_index, value=value)

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 4, 28)

    file_path = os.path.join(EXPORTS_DIR, f"marksheet_{sheet_id}.xlsx")
    workbook.save(file_path)
    return file_path
